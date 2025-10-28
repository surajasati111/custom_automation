import requests
import pandas as pd
import dateutil.relativedelta
import json
import urllib3
import logging
import os
import openpyxl
import shutil
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import argparse

parser = argparse.ArgumentParser(description='Optional app description')

# Optional positional argument
parser.add_argument('month', type=int, nargs='?',
                    help='Please enter the month name argument like 05')
                    
args = parser.parse_args()


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

dir_name = os.path.dirname(__file__)
path = f"{dir_name}/"
today_date = datetime.now().date()
now_date_time = datetime.now()

file_name = f'{dir_name}/psirt_kpi_{today_date}.xlsx'
shutil.copy(f'{dir_name}/kpi_column.xlsx', file_name)

log_path = f'{dir_name}/logs/{today_date}_psirt_incident.log'
logging.basicConfig(filename=log_path,
                    level=logging.DEBUG,
                    format='%(asctime)s:%(levelname)s:%(name)s:%(message)s',
                    datefmt='%d-%b-%y %H:%M:%S')


def convert_date(date_):
    return str(date_.strftime('%Y-%m-%dT' + '00:00:00.0000000Z'))


def convert_date_start(date_):
    return str(date_.strftime('%Y-%m-%dT' + '00:00:00 +0530'))


def convert_date_end(date_):
    return str(date_.strftime('%Y-%m-%dT' + '23:59:59 +0530'))


def get_xsoar_data(qry_type):
    url_ = ''  # Global variable for managing the instance URL
    apikey = ''  # key for prod
    #apikey = ''  # key for prod
    
    # url_ = ""
    # apikey = ''  # key for prod

    # server url for the API call
    url = f"{url_}/incidents/search"

    page = 0  # page number ot start at
    size = 500  # number of Incidents to return per page
    curcuitbreaker = 3
    
    date_format = "%m/%d/%Y"
    current_dt = datetime.strptime(f'{datetime.now().month}/{datetime.now().day}/{datetime.now().year}', date_format)
    b = datetime.strptime(get_current_month_last_date(), date_format)
    delta = current_dt - b
    print("diff=", delta.days) 

    from_date = dateutil.relativedelta.relativedelta(days=364 + delta.days)
    to_date = dateutil.relativedelta.relativedelta(days=delta.days)
    
    current_date = datetime.strptime(str(today_date), "%Y-%m-%d")
    
    date_range = False
    if qry_type == "all":
        from_date = convert_date_start(current_date - from_date)
        to_date = convert_date_end(current_date - to_date)
        query = f'type:"PSIRT Rule" -vrepid:"" -firstresponse:"" (occurred:>="{from_date}"  and occurred:<="{to_date}")'
    elif qry_type == "no response":
        query = 'type:"PSIRT Rule" -vrepid:"" firstresponse:""'
        from_date = convert_date_start(current_date - from_date)
        to_date = convert_date_end(current_date - to_date)
    elif qry_type == "active":
        from_date = dateutil.relativedelta.relativedelta(days=600)
        to_date = dateutil.relativedelta.relativedelta(days=delta.days)
        to_date = convert_date_end(current_date - to_date)
        query = f'type:"PSIRT Rule" -vrepid:"" status:Active occurred:<="{to_date}"'
    else:
        # For closed or Slide 3
        adv_date_to_date = convert_date_end(current_date - to_date)       
        query = f'type:"PSIRT Rule" -vrepid:"" (advisoryreleasedate:>="{convert_date_start(current_date - from_date)}" and advisoryreleasedate:<="{adv_date_to_date}") (fixreleasedate:<="{adv_date_to_date}" or fixreleasedate="") closed:<="{adv_date_to_date}"'
        date_range = True
    
    print("query=================", query)
    logging.info(f"from date={from_date} to To Date={to_date}")
    print(f"from date={from_date} to To Date={to_date}")
    
    # headers for request
    headers = {
        "Authorization": apikey,
        "accept": "application/json",
        "content-type": "application/json"
    }

    logging.info(f"query= {query}")
    body = {
        "userFilter": False,
        "filter": {
            "page": page,
            "size": size,
            "query": f"{query}",
        }
    }

    # get incidents
    res = requests.post(f"{url}", headers=headers, json=body, verify=False)
    total = res.json().get('total', 0)
    incidents = res.json().get('data', [])
    count = 1

    # if there are more events than the default size, page through and get them all
    while len(incidents) < total:
        body['filter']['page'] = count
        res = requests.post(f"{url}", headers=headers, json=body, verify=False)
        incidents.extend(res.json().get('data', []))
        count += 1
        # break if this goes crazy
        if count == curcuitbreaker:
            break

    print(f"Incidents Found: {len(incidents)}, Total Incidents: {total}, Number of API calls made:Number of API calls made: {count}")
    logging.info(f"Incidents Found: {len(incidents)} , Total Incidents: {total}, Number of API calls made: {count} ")
    
    if date_range:
        # headers for request
        headers = {
            "Authorization": apikey,
            "accept": "application/json",
            "content-type": "application/json"
        }
        print(f'(id:362281 or id:362266 or id:362261 or id:362257 or id:362251 or id:357447) and (advisoryreleasedate:>="{convert_date_start(current_date - from_date)}"  and advisoryreleasedate:<="{convert_date_end(current_date - to_date)}")')


        body = {
            "userFilter": False,
            "filter": {
                "page": page,
                "size": size,
                #"query": "id:362281 or id:362266 or id:362261 or id:362257 or id:362251 or id:357447",
                "query": f'(id:362281 or id:362266 or id:362261 or id:362257 or id:362251 or id:357447) and (advisoryreleasedate:>="{convert_date_start(current_date - from_date)}"  and advisoryreleasedate:<="{to_date}")',
            }
        }

        # get incidents
        res = requests.post(f"{url}", headers=headers, json=body, verify=False)
        total = res.json().get('total', 0)
        incidents_2 = res.json().get('data', [])
        count = 1

        # if there are more events than the default size, page through and get them all
        while len(incidents_2) < total:
            body['filter']['page'] = count
            res = requests.post(f"{url}", headers=headers, json=body, verify=False)
            incidents_2.extend(res.json().get('data', []))
            count += 1
            # break if this goes crazy
            if count == curcuitbreaker:
                break  

        return incidents_2 + incidents # Create a copy to avoid modifying data1 directly
    return incidents


def change_date_format(date_):
    if "T" in date_:
        new_date = date_.split("T")
        new_date_f = new_date[0].split("-")
        new_time_f = new_date[1].split(":")
        return f"{new_date_f[1]}/{new_date_f[2]}/{new_date_f[0]} {new_time_f[0]}:{new_time_f[1]}"
    else:
        space_count = date_.count(" ")
        new_date = date_.split("  ") if space_count >= 3 else date_.split(" ")
        new_time_f = new_date[1].split(":")
        return f"{new_date[0]} {new_time_f[0]}:{new_time_f[1]}"


def get_severity(sev):
    match sev:
        case 1:
            return "Low"
        case 2:
            return "Medium"
        case 3:
            return "High"
        case 4:
            return "Critical"
        case _:
            return "Unknown"
            

def get_days_as_vh(sev, customer="Unknown"):
    # print(sev, customer)
    if sev in ["High", "Critical"]:
        vh = 14
    elif sev in "Medium":
        vh = 28
    else:
        vh = 0
    return vh
    

def get_days_as_vh_old(sev, customer="Unknown"):
    # print(sev, customer)
    if sev in ["High", "Critical"] and customer in "Known":
        vh = 182
    elif sev in "Medium" and customer in "Known":
        vh = 364
    elif sev in ["High", "Critical"] and customer in "Unknown":
        vh = 14
    elif sev in "Medium" and customer in "Unknown":
        vh = 28
    else:
        vh = 0
    return vh


def get_days_diff(date1, date2):
    if len(date1) < 3 or len(date2) < 3:
        return 0

    date_format = "%m/%d/%Y"

    new_date1 = date1.split(" ")
    new_date2 = date2.split(" ")
    a = datetime.strptime(new_date1[0], date_format)
    b = datetime.strptime(new_date2[0], date_format)
    delta = b - a
    #print("days_dif=", delta.days, type(delta.days))
    #exit()
    return 0 if delta.days < 0 else delta.days 


def get_sadv(no_of_days, ideals_days_as_vh):
    flag = "No"
    if no_of_days <= ideals_days_as_vh:
        flag = "Yes"
    return flag


def get_hours(start_, end_):
    return (datetime.strptime(end_, "%H:%M") - datetime.strptime(start_, "%H:%M")).seconds // 3600


def get_current_month_last_date():
    input_dt = datetime(datetime.now().year, args.month , datetime.now().day)
    next_month = input_dt.replace(day=28) + timedelta(days=4)
    res = next_month - timedelta(days=next_month.day)
    print((res.date()).strftime('%m/%d/%Y 12:00'))
    return ((res.date()).strftime('%m/%d/%Y'))


def active_incident(sheet_, data_, ignored_list=[]):
    row_ = 2
    column_ = 0
    for item in data_:
        if item["CustomFields"]["vrepid"] not in ignored_list:
            created_date = change_date_format(item["CustomFields"]["reporteddate"]) if item["CustomFields"].get(
                "reporteddate") is not None else change_date_format(item["occurred"])
            adv_date = get_current_month_last_date()
            adv_date = f'{adv_date} 12:00'
            print(adv_date, "--", item["id"], "--", item["severity"])
            get_customer = item["CustomFields"]["affectedproductcustomerbase"] if item["CustomFields"].get(
                "affectedproductcustomerbase") is not None else "Known"
            get_sev = get_severity(item["severity"])
            no_days = get_days_diff(created_date, adv_date)
            idel_as_vh = get_days_as_vh(get_sev, get_customer)
            # Or, access by row and column indices
            sheet_.cell(row=row_, column=column_ + 1, value=row_-1)
            sheet_.cell(row=row_, column=column_ + 2, value=item["CustomFields"]["vrepid"])
            sheet_.cell(row=row_, column=column_ + 3,
                        value=item["CustomFields"]["businessarea"] if item["CustomFields"].get(
                            "businessarea") is not None else "")
            sheet_.cell(row=row_, column=column_ + 4,
                        value=item["CustomFields"]["businessdivision"] if item["CustomFields"].get(
                            "businessdivision") is not None else "")
            sheet_.cell(row=row_, column=column_ + 5, value="In Progress" if item["status"] == 1 else "Closed")
            sheet_.cell(row=row_, column=column_ + 6, value=created_date)
            sheet_.cell(row=row_, column=column_ + 7, value=adv_date)
            sheet_.cell(row=row_, column=column_ + 8, value=no_days)
            sheet_.cell(row=row_, column=column_ + 9, value=item["CustomFields"]["reportedemailfrom"])
            sheet_.cell(row=row_, column=column_ + 10,
                        value=item["CustomFields"]["advisoryreleasedate"] if item["CustomFields"].get(
                            "advisoryreleasedate") is not None else "NA")
            sheet_.cell(row=row_, column=column_ + 11,
                        value=item["CustomFields"]["advisoryrelease"] if item["CustomFields"].get(
                            "advisoryrelease") is not None else "Not Due")
            sheet_.cell(row=row_, column=column_ + 12,
                        value=item["CustomFields"]["fixrelease"] if item["CustomFields"].get(
                            "fixrelease") is not None else "Not Yet")
            row_ += 1


def closed_incident(sheet, data_, colmn=None, ignored_list=[]):
    # Update a cell value
    row_ = 2
    column_ = 0

    for item in data_:
        if item["CustomFields"]["vrepid"] not in ignored_list:
            print("inc_id=",item["id"])
            created_date = change_date_format(item["CustomFields"]["reporteddate"]) if item["CustomFields"].get(
                "reporteddate") is not None else change_date_format(item["occurred"])
            
            adv_date = change_date_format(item["CustomFields"]["advisoryreleasedate"]) if item["CustomFields"].get(
                "advisoryreleasedate") is not None else "NA"
            fix_date = change_date_format(item["CustomFields"]["fixreleasedate"]) if item["CustomFields"].get(
                "fixreleasedate") is not None else ""
            print(adv_date, "--", fix_date, item["id"], "--", item["severity"])
            get_days_diff(adv_date, fix_date)
            get_custmer = item["CustomFields"]["affectedproductcustomerbase"] if item["CustomFields"].get(
                "affectedproductcustomerbase") is not None else "Known"
            get_sev = get_severity(item["severity"])
            no_days = get_days_diff(fix_date, adv_date)
            idel_as_vh = get_days_as_vh(get_sev, get_custmer)
            sadv = get_sadv(no_days, idel_as_vh)
            get_n = no_days if no_days >= 0 else 0
            sheet.cell(row=row_, column=column_ + 1, value=row_-1)
            sheet.cell(row=row_, column=column_ + 2, value="In Progress" if item["status"] == 1 else "Closed")
            sheet.cell(row=row_, column=column_ + 3, value=item["CustomFields"]["vrepid"])
            sheet.cell(row=row_, column=column_ + 4, value=item["CustomFields"]["businessarea"] if item["CustomFields"].get(
                "businessarea") is not None else "")
            sheet.cell(row=row_, column=column_ + 5,
                       value=item["CustomFields"]["businessdivision"] if item["CustomFields"].get(
                           "businessdivision") is not None else "")
            sheet.cell(row=row_, column=column_ + 6, value=created_date)
            sheet.cell(row=row_, column=column_ + 7, value=adv_date)
            sheet.cell(row=row_, column=column_ + 8, value=fix_date)
            sheet.cell(row=row_, column=column_ + 9, value=get_n)
            sheet.cell(row=row_, column=column_ + 10, value=idel_as_vh)
            sheet.cell(row=row_, column=column_ + 11, value="Yes" if item["CustomFields"].get("informedinternalcustomer") is not None and item["CustomFields"].get("informedinternalcustomer") == "Yes" else sadv)
            sheet.cell(row=row_, column=column_ + 12, value=get_days_diff(created_date, fix_date))
            sheet.cell(row=row_, column=column_ + 13, value=item["CustomFields"]["vulnerabilityidentified"] if item["CustomFields"].get(
                           "vulnerabilityidentified") is not None else "")

            #if colmn:
            sheet.cell(row=row_, column=column_ + 14,
                       value=item["CustomFields"]["fixrelease"] if item["CustomFields"].get(
                           "fixrelease") is not None else "No")

            sheet.cell(row=row_, column=column_ + 15,
                       value=item["CustomFields"]["advisoryrelease"] if item["CustomFields"].get(
                           "advisoryrelease") is not None else "No")
            sheet.cell(row=row_, column=column_ + 16, value=get_sev)
            sheet.cell(row=row_, column=column_ + 17, value=get_custmer)
            sheet.cell(row=row_, column=column_ + 18, value=item["closeNotes"])
            row_ += 1
    

        # to add hard code entry for 
        sheet.cell(row=row_, column=column_ + 1, value=row_-1)
        sheet.cell(row=row_, column=column_ + 2, value="Closed")
        sheet.cell(row=row_, column=column_ + 3, value="ABBVREP0188")
        sheet.cell(row=row_, column=column_ + 4, value="EL")
        sheet.cell(row=row_, column=column_ + 5, value="ELSP")
        sheet.cell(row=row_, column=column_ + 6, value="12/16/2024 12:00")
        sheet.cell(row=row_, column=column_ + 7, value="06/26/2025 12:46")
        sheet.cell(row=row_, column=column_ + 8, value="04/10/2025 12:46")
        sheet.cell(row=row_, column=column_ + 9, value=77)
        sheet.cell(row=row_, column=column_ + 10, value=28)
        sheet.cell(row=row_, column=column_ + 11, value="No")
        sheet.cell(row=row_, column=column_ + 12, value=192)
        sheet.cell(row=row_, column=column_ + 13, value="Internal")

        if colmn:
            sheet.cell(row=row_, column=column_ + 14, value="Fixed")
            sheet.cell(row=row_, column=column_ + 15, value="Yes")
            sheet.cell(row=row_, column=column_ + 16, value="Medium")
            sheet.cell(row=row_, column=column_ + 17, value="Unknown")
            sheet.cell(row=row_, column=column_ + 18, value="PSIRT published the advisory,  CVE also published, email notification sent, all the actions are taken. Hence closing the case.")


def no_response_incident(sheet_, incidents):
    row_ = 2
    column_ = 0

    for item in incidents:
        print("inc_id=", item["id"])
        create_date = change_date_format(item["CustomFields"]["createdtime"]) if item["CustomFields"].get(
            "createdtime") is not None else change_date_format(item["occurred"])
        # Or, access by row and column indices
        sheet_.cell(row=row_, column=column_ + 1, value=item["CustomFields"]["vrepid"])
        sheet_.cell(row=row_, column=column_ + 2, value=create_date)
        sheet_.cell(row=row_, column=column_ + 4,
                    value=item["CustomFields"]["businessarea"] if item["CustomFields"].get(
                        "businessarea") is not None else "")
        sheet_.cell(row=row_, column=column_ + 5,
                    value=item["CustomFields"]["businessdivision"] if item["CustomFields"].get(
                        "businessdivision") is not None else "")
        row_ += 1


def first_response_incident(sheet_, incidents, ignored_list=[]):
    row_ = 2
    column_ = 0

    for item in incidents:
        print("inc_id=", item["id"])
        create_date = change_date_format(item["CustomFields"]["reporteddate"]) if item["CustomFields"].get(
            "createdtime") is not None else change_date_format(item["occurred"])
        first_response_date = change_date_format(item["CustomFields"]["firstresponse"])
        # print(create_date, "----", first_response_date)
        get_diff = get_days_diff(create_date, first_response_date)
        crete_time = create_date.split(" ") if get_diff > 0 else 0
        first_time = first_response_date.split(" ") if get_diff > 0 else 0
        #print(crete_time, "++", first_time)
        final_c = str("00:00" if crete_time == 0 else crete_time[1])
        final_f = str("00:00" if first_time == 0 else first_time[1])
        get_hours_ = get_hours(final_c, final_f)
        # Or, access by row and column indices
        sheet_.cell(row=row_, column=column_ + 1, value=item["CustomFields"]["vrepid"])
        sheet_.cell(row=row_, column=column_ + 2, value=create_date)
        sheet_.cell(row=row_, column=column_ + 3, value=first_response_date)
        sheet_.cell(row=row_, column=column_ + 4,
                    value=item["CustomFields"]["businessarea"] if item["CustomFields"].get(
                        "businessarea") is not None else "")
        sheet_.cell(row=row_, column=column_ + 5,
                    value=item["CustomFields"]["businessdivision"] if item["CustomFields"].get(
                        "businessdivision") is not None else "")
        sheet_.cell(row=row_, column=column_ + 6, value=f"{get_diff}.{get_hours_}")
        sheet_.cell(row=row_, column=column_ + 7, value="Yes" if get_diff <= 2 or item["CustomFields"]["vrepid"] in ignored_list else "No")
        row_ += 1


def read_write_summary(file_, sheet_name, column_filter, type_=None):
    df = pd.read_excel(file_, engine='openpyxl', sheet_name=sheet_name)
    # Get the count of 'Yes' and 'No' based on Division
    division_counts = df.groupby(column_filter).size().unstack(fill_value=0)
    # Extract the division_counts variable
    division_counts_dict = division_counts.to_dict()

    # Initialize the new dictionary
    updated_dict = {}
    # Iterate over the original dictionary to transform it
    for key, value in division_counts_dict.items():
        for sub_key, sub_value in value.items():
            if sub_key not in updated_dict:
                updated_dict[sub_key] = {}
            updated_dict[sub_key][key] = sub_value

    final_dict = {}

    if type_ == "ND":
        for item in updated_dict.items():
            if item[1].get("due"):
                # print(due, (item[1]["Not due"]*100)/(due + item[1]["Not due"]))
                final_dict[item[0]] = round((item[1]["Not due"] * 100) / (item[1]["due"] + item[1]["Not due"]), 2)
            else:
                final_dict[item[0]] = 100
    else:
        for item in updated_dict.items():
            if item[1].get("No"):
                print(item[1]["No"], (item[1]["Yes"] * 100) / (item[1]["No"] + item[1]["Yes"]))
                final_dict[item[0]] = round((item[1]["Yes"] * 100) / (item[1]["No"] + item[1]["Yes"]), 2)
            else:
                final_dict[item[0]] = 100
    return final_dict


def read_summary_cells(dataframe1):
    create_new_dict = {}
    # Iterate the loop to read the cell values
    for row in range(3, dataframe1.max_row + 1):
        # print(dataframe1[row][1].value, row)
        create_new_dict[dataframe1[row][1].value] = row

    print(create_new_dict)
    return create_new_dict


def get_color_day(days):
    if days in [0, 'n/a']:
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF")
    elif days > 365:
        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif 180 < days < 365:
        fill = PatternFill(start_color="FFC104", end_color="FFC104", fill_type="solid")
    else:
        fill = PatternFill(start_color="92D24F", end_color="92D24F", fill_type="solid")
    return fill


def get_color_percent(time_):
    if time_ in [0, 'n/a']:
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF")
    elif time_ <= 59:
        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif 60 <= time_ <= 89:
        fill = PatternFill(start_color="FFC104", end_color="FFC104", fill_type="solid")
    else:
        fill = PatternFill(start_color="92D24F", end_color="92D24F", fill_type="solid")
    return fill


def summary_kpi_update(dataframe1, create_new_dict, final_dict, column_update, days=None):
    row_ = 3
    column_ = 0

    for key, val in create_new_dict.items():
        # print(row_, column, value=final_dict.get(key, 0))
        colmn_val = int(round(final_dict.get(key))) if final_dict.get(key) else "n/a"
        dataframe1.cell(row=row_, column=column_update, value=colmn_val)
        if days:
            dataframe1.cell(row=row_, column=column_update, value=colmn_val).fill = get_color_day(colmn_val)
        else:
            dataframe1.cell(row=row_, column=column_update, value=f"{int(round(final_dict.get(key)))}%" if final_dict.get(key) is not None else "n/a").fill = get_color_percent(colmn_val)
        print("Final Cell value=", colmn_val)
        row_ += 1
        if row_ >= 24:
            break
        # print("this is the row num=", row_)
        

def summary_kpi_update_2(dataframe1, create_new_dict, final_dict, column_update):
    row_ = 3
    column_ = 0

    for key, val in create_new_dict.items():
        # print(row_, column, value=final_dict.get(key, 0))
        colmn_val = int(round(final_dict.get(key))) if final_dict.get(key) else "n/a"
        dataframe1.cell(row=row_, column=column_update, value=f"{int(round(final_dict.get(key)))}%" if final_dict.get(key) is not None else "n/a")
        row_ += 1
        if row_ >= 24:
            break


def number_calculation(file_, sheet_name, comun_, operation="avg"):
    df = pd.read_excel(file_, engine='openpyxl', sheet_name=sheet_name)
    df[comun_[1]] = pd.to_numeric(df[comun_[1]], errors='coerce')
    # Calculate the average based on Division, skipping NaN values in the calculation
    if operation == "avg":
        by_division = df.groupby(comun_[0])[comun_[1]].mean()
    else:
        by_division = df.groupby(comun_[0])[comun_[1]].max()

    return by_division.to_dict()


def mail_send():
    import win32com.client as win32
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = "harmeet.singh@in.abb.com;suryaprakash.asti@in.abb.com;arun.kumarrai@in.abb.com"
    mail.Subject = 'PSIRT KPI Report'
    mail.Body = 'Check out the attached Report..'
    mail.HTMLBody = '<h2>Check out the attached Report.</h2>' #this field is optional
    # To attach a file to the email (optional):
    attachment  = file_name
    mail.Attachments.Add(attachment)
    mail.Send()


if __name__ == '__main__':
    logging.info("*************************")
    logging.info(f"Script started at {now_date_time}")
    
    if args.month:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_name)

        # for all the incidents
        all_incident = get_xsoar_data("all")
        if all_incident:
            first_response_incident(workbook['KPI_1'], all_incident, ignored_list=["ABBVREP0186", "ABBVREP0201"])

        # for KPI 2 incidents
        all_incident = get_xsoar_data("no response")
        if all_incident:
            no_response_incident(workbook['KPI_2'], all_incident)

        # for active incidents
        all_incident = get_xsoar_data("active")
        if all_incident:
            active_incident(workbook['KPI_4'], all_incident, ignored_list=[])
            active_incident(workbook['KPI_7'], all_incident, ignored_list=[])
            active_incident(workbook['KPI_8'], all_incident, ignored_list=[])
        
        # for closed incidents
        all_incident = get_xsoar_data("closed")
        if all_incident:
            closed_incident(workbook['KPI_3'], all_incident, ignored_list=["ABBVREP0175"])
            closed_incident(workbook['KPI_5'], all_incident, colmn="13", ignored_list=["ABBVREP0175"])
            closed_incident(workbook['KPI_6'], all_incident, ignored_list=["ABBVREP0175"])

        # workbook.save("kpi.xlsx")
        
        workbook.save(file_name)

        import time

        time.sleep(10)

        workbook = openpyxl.load_workbook(file_name)

        kpi_1_summary = read_write_summary(file_name, "KPI_1", ['ABB Division', 'Sack'])
        wb = workbook["KPI Summary"]
        summary_kpi_update(wb, read_summary_cells(wb), kpi_1_summary, 5)

        # update KPI_2
        kpi_2_summary = read_write_summary(file_name, "KPI_2", ['ABB Division', 'ABB Business Area'])
        wb = workbook["KPI Summary"]
        summary_kpi_update_2(wb, read_summary_cells(wb), kpi_2_summary, 6)
        
        # update KPI_3
        kpi_3_summary = read_write_summary(file_name, "KPI_3", ['ABB Division', 'Sadv'])
        summary_kpi_update(wb, read_summary_cells(wb), kpi_3_summary, 7)

        # update KPI_4
        kpi_4_summary = read_write_summary(file_name, "KPI_4", ['ABB Division', 'Advisory due status'], "ND")
        summary_kpi_update(wb, read_summary_cells(wb), kpi_4_summary, 8)

        # update KPI_5
        kpi_5_summary = number_calculation(file_name, "KPI_5", ['ABB Division', 'tf-tr(time of fix - time of report)'],
                                           "avg")
        summary_kpi_update(wb, read_summary_cells(wb), kpi_5_summary, 9, "Yes")

        # update KPI_6
        kpi_6_summary = number_calculation(file_name, "KPI_6", ['ABB Division', 'tf-tr(time of fix - time of report)'],
                                           "max")
        summary_kpi_update(wb, read_summary_cells(wb), kpi_6_summary, 10, "Yes")

        # update KPI_7
        kpi_7_summary = number_calculation(file_name, "KPI_7", ['ABB Division', 'Tnow-Tr'], "avg")
        summary_kpi_update(wb, read_summary_cells(wb), kpi_7_summary, 11, "Yes")

        # update KPI_8
        kpi_8_summary = number_calculation(file_name, "KPI_8", ['ABB Division', 'Tnow-Tr'], "max")
        summary_kpi_update(wb, read_summary_cells(wb), kpi_8_summary, 12, "Yes")
        workbook.save(file_name)

        #mail_send()
    else:
        print("please enter the month number like 05 for May, 06 for June")
    logging.info(f"Script ended at {datetime.now()}")
    
